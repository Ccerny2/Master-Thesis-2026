"""
Follow-up verification for the 57 rows that couldn't be verified in the first pass.
Handles Year Planned, news pages, mismatches, and annotates edge cases.
"""

import requests
from bs4 import BeautifulSoup
import re
import time
import openpyxl
from urllib.parse import urlparse

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
        'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    )
}
TIMEOUT = 20
DELAY = 0.6

XLSX = '/Users/elenamurray/Documents/Documents/Repositories/data_journalism/data_bit_1/va_pnnl_researched.xlsx'
wb = openpyxl.load_workbook(XLSX)
ws = wb.active

existing_headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
col_verified  = existing_headers['verified']
col_src_year  = existing_headers['source_year_found']
col_src_field = existing_headers['source_year_field']
col_note      = existing_headers['verification_note']
col_new_src   = existing_headers['new_source_url']

COL_YEAR = 14
COL_BASIS = 15
COL_URL  = 16


# ── Extraction helpers ──────────────────────────────────────────────────────

def extract_baxtel_extended(raw_html):
    """Return (year, field) — covers Year Built, Year Planned, Year Opened."""
    for label in ('Year Built', 'Year Opened', 'Year Planned'):
        m = re.search(rf'{label}:\s*(\d{{4}})', raw_html)
        if m:
            return m.group(1), label
    m = re.search(r'[Cc]onstructed in (\d{4})', raw_html)
    if m:
        return m.group(1), 'Constructed in (text)'
    return None, None


def extract_datacenter_fyi(raw_html):
    m = re.search(
        r'Operational Date</p>.*?(\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{1,2},?\s+(\d{4})|\b(\d{4})\b)',
        raw_html, re.DOTALL
    )
    if m:
        return m.group(2) or m.group(3), 'Operational Date'
    return None, None


def extract_baxtel_news(raw_html, research_year):
    """For baxtel news pages: look for the research year in the press release date or body."""
    m = re.search(rf'\b{research_year}\b', raw_html)
    if m:
        return str(research_year), 'Year mentioned in news article'
    return None, None


def extract_generic(raw_html, soup):
    text = soup.get_text(' ', strip=True)
    m = re.search(r'(?:Year (?:Built|Opened|Planned)|Opened|Built)[:\s]+(\d{4})', text, re.IGNORECASE)
    if m:
        return m.group(1), m.group(0).split(':')[0].strip()
    return None, None


cache = {}
last_req = {}

def fetch(url):
    if url in cache:
        return cache[url]
    domain = urlparse(url).netloc
    elapsed = time.time() - last_req.get(domain, 0)
    if elapsed < DELAY:
        time.sleep(DELAY - elapsed)
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        cache[url] = (r.status_code, r.text)
    except Exception as e:
        cache[url] = (0, str(e))
    last_req[domain] = time.time()
    return cache[url]


# ── Manual overrides & notes from investigation ────────────────────────────
# Keyed by row number → (verified, src_year, src_field, note, new_src)

MANUAL = {
    # CoreSite VA3: Year Built 2019 (original); research year 2025 = new 12k sqft
    # data hall mentioned on same source page. Partially verifiable.
    18: ('Partial - see note', '2019', 'Year Built (original building)',
         'Baxtel shows Year Built 2019 for original VA3 building. '
         'Research year 2025 refers to a new 12,000 sq ft data hall '
         'completed in 2025 per same baxtel page: '
         '"CoreSite completed construction of a new 38,000 sq ft data hall '
         'suite at VA3 in Nov 2024; another 12,000 sq ft data hall scheduled '
         'for completion in 2025."', ''),

    301: ('Partial - see note', '2019', 'Year Built (original building)',
          'Same source as CoreSite VA3 (row 18). Source shows Year Built 2019 '
          'for original building. Research year 2025 refers to a new data hall '
          'suite on same campus confirmed in source narrative.', ''),

    # Amazon IAD128: source page covers IAD-124/125/126/127 (Dec 2023) — not IAD-128.
    # Source is misapplied.
    166: ('Could not verify - source mismatch', '2023', 'Operational Date (IAD-124-127)',
          'Source page is for IAD-124/125/126/127 (operational Dec 2023), '
          'not IAD-128. Year 2024 for IAD-128 cannot be verified from this source.', ''),

    # Amazon IAD601: same misapplied source
    208: ('Could not verify - source mismatch', '2023', 'Operational Date (IAD-124-127)',
          'Source page covers IAD-124/125/126/127 (Dec 2023), not IAD-601. '
          'Year 2024 for IAD-601 cannot be confirmed from this source.', ''),

    # Iron Mountain VA-6: Baxtel shows Year Built 2025, research says 2024.
    280: ('Mismatch', '2025', 'Year Built',
          'Baxtel shows Year Built: 2025 but research year is 2024. '
          'Construction may have been recorded at completion in 2025.', ''),

    # Equinix DC97: source is baxtel news from May 2017 — acquisition 2017 confirmed
    20: ('Verified', '2017', 'Press release date (May 2017)',
         'Baxtel news article dated May 2017 confirms Equinix acquisition '
         'of 29 Verizon data centers including DC97 in 2017.', ''),

    # Element Critical VA1: page says "acquired in late 2016"
    35: ('Verified', '2016', 'Acquisition year (text)',
         'Baxtel page states Element Critical acquired VA1 in late 2016.', ''),
    63: ('Verified', '2016', 'Acquisition year (text)',
         'Baxtel page states Element Critical acquired VA1 in late 2016.', ''),

    # BlackChamber / Arcola Tech Park: news article from 2022, not 2024
    168: ('Could not verify', '2022', 'News article publication year',
          'Baxtel news source is from October 2022 and covers BlackChamber '
          'announcing plans for Arcola. Research year 2024 not mentioned '
          'in source.', ''),
    284: ('Could not verify', '2022', 'News article publication year',
          'Same baxtel news article (Oct 2022). Research year 2024 not '
          'mentioned in source.', ''),

    # NTT Ashburn Campus: no individual year — campus-level page only
    102: ('Could not verify', None, None,
          'Baxtel shows campus-level page with no Year Built/Planned. '
          'Research year 2017 not confirmed.', ''),
    238: ('Could not verify', None, None,
          'Same NTT Ashburn Campus page. No year field present.', ''),

    # CoreSite category page (not individual VA2 page)
    223: ('Could not verify', None, None,
          'Source URL is CoreSite company page, not individual VA2 page. '
          'No year found for VA2 specifically.', ''),

    # Digital Realty Loudoun Ashburn Campus: no year field; campus is operational
    # All rows 55,57,58,59,96-99,121,122,147,200,201,206,283 same URL
    55:  ('Could not verify', None, None,
          'Digital Realty Loudoun Ashburn Campus page shows facility as '
          'Operational but no Year Built/Planned field. 2026 not confirmed.', ''),
    57:  ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    58:  ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    59:  ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    96:  ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    97:  ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    98:  ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    99:  ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    121: ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    122: ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    147: ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    200: ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    201: ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    206: ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),
    283: ('Could not verify', None, None,
          'Same Digital Realty campus page. No year field. 2026 not confirmed.', ''),

    # TA Realty Leesburg (True North): page has no explicit completion date
    171: ('Could not verify', None, None,
          'Baxtel page for TA Realty Leesburg has no Year Built/Planned. '
          '2027 not confirmed from source.', ''),
    275: ('Could not verify', None, None,
          'Same TA Realty Leesburg page. No year field. 2027 not confirmed.', ''),
    276: ('Could not verify', None, None,
          'Same TA Realty Leesburg page. No year field. 2027 not confirmed.', ''),
    277: ('Could not verify', None, None,
          'Same TA Realty Leesburg page. No year field. 2027 not confirmed.', ''),
    298: ('Could not verify', None, None,
          'Same TA Realty Leesburg page. No year field. 2027 not confirmed.', ''),
    299: ('Could not verify', None, None,
          'Same TA Realty Leesburg page. No year field. 2027 not confirmed.', ''),

    # Lumos Lynchburg: no year on baxtel
    314: ('Could not verify', None, None,
          'Baxtel page for Lumos Lynchburg has no year information. '
          '2017 not confirmed.', ''),
}

# Rows using baxtel.com/news/* pages that aren't in MANUAL yet
NEWS_ROWS = {20}  # already handled above

# ── Rows that need re-extraction with Year Planned support ─────────────────
RECHECK_ROWS = [3, 22, 28, 29, 33, 34, 139, 151, 164, 210, 247, 306, 319, 320, 253, 268, 310]
# Plus COPT PDF rows 22, 210 and 404 rows 16, 29, 164

FAILED_ROWS = [2, 3, 16, 18, 20, 22, 28, 29, 33, 34, 35, 55, 57, 58, 59, 63, 96, 97, 98,
               99, 102, 121, 122, 139, 146, 147, 151, 164, 166, 168, 171, 200, 201, 206,
               207, 208, 210, 222, 223, 238, 247, 253, 268, 275, 276, 277, 280, 283, 284,
               298, 299, 301, 306, 310, 314, 319, 320]

print('Processing follow-up rows...\n')

for row in FAILED_ROWS:
    research_year = ws.cell(row, COL_YEAR).value
    basis         = ws.cell(row, COL_BASIS).value
    url           = ws.cell(row, COL_URL).value
    name          = ws.cell(row, 9).value
    operator      = ws.cell(row, 7).value

    # Apply manual overrides
    if row in MANUAL:
        verified, src_year, src_field, note, new_src = MANUAL[row]
        ws.cell(row, col_verified).value  = verified
        ws.cell(row, col_src_year).value  = src_year
        ws.cell(row, col_src_field).value = src_field
        ws.cell(row, col_note).value      = note
        if new_src:
            ws.cell(row, col_new_src).value = new_src
        print(f'Row {row:3d}: [MANUAL] {verified}  | {name}')
        continue

    if not url:
        print(f'Row {row:3d}: no URL')
        continue

    # Attempt re-fetch with extended extraction
    status, raw = fetch(url)

    domain = urlparse(url).netloc
    soup = BeautifulSoup(raw, 'html.parser') if status == 200 else None

    if status == 0 or status == 404:
        ws.cell(row, col_verified).value = 'Could not verify'
        ws.cell(row, col_note).value = f'HTTP {status} - source URL unavailable'
        print(f'Row {row:3d}: HTTP {status}  | {name}  | {url}')
        continue

    if status != 200:
        ws.cell(row, col_verified).value = 'Could not verify'
        ws.cell(row, col_note).value = f'HTTP {status}'
        print(f'Row {row:3d}: HTTP {status}  | {name}')
        continue

    # Extract year
    if '/news/' in url:
        src_year, src_field = extract_baxtel_news(raw, research_year)
    elif 'baxtel.com' in domain or 'cdn.baxtel.com' in domain:
        src_year, src_field = extract_baxtel_extended(raw)
    elif 'datacenter.fyi' in domain:
        src_year, src_field = extract_datacenter_fyi(raw)
    elif 'cleanview.co' in domain:
        src_year, src_field = None, None  # JS-rendered, can't extract
    else:
        src_year, src_field = extract_generic(raw, soup)

    ws.cell(row, col_src_year).value  = src_year
    ws.cell(row, col_src_field).value = src_field

    if not src_year:
        ws.cell(row, col_verified).value = 'Could not verify'
        ws.cell(row, col_note).value = (
            'No year field found on source page. '
            f'Research year {research_year} ({basis}) cannot be confirmed.'
        )
        print(f'Row {row:3d}: no year extracted  | {name}')
    elif src_year == str(research_year):
        ws.cell(row, col_verified).value = 'Verified'
        ws.cell(row, col_note).value = f'Source shows {src_field}: {src_year}'
        print(f'Row {row:3d}: ✓ {research_year}  | {name}')
    else:
        ws.cell(row, col_verified).value = 'Mismatch / Could not verify'
        ws.cell(row, col_note).value = (
            f'Research year {research_year} ({basis}); '
            f'source shows {src_field}: {src_year}'
        )
        print(f'Row {row:3d}: ✗ research={research_year} vs source={src_year}  | {name}')

wb.save(XLSX)
print('\nDone. Saved.')

# Print summary
still_unverified = []
for row in range(2, ws.max_row + 1):
    v = ws.cell(row, col_verified).value
    if v and 'Could not verify' in str(v):
        still_unverified.append(row)

print(f'\nRows still unverified: {len(still_unverified)}')
for r in still_unverified:
    print(f'  Row {r:3d}: {ws.cell(r,9).value} | year={ws.cell(r,14).value} | {ws.cell(r,16).value}')
