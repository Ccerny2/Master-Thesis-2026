"""
Final verification pass — applies researched conclusions to the remaining rows.
"""

import openpyxl

XLSX = '/Users/elenamurray/Documents/Documents/Repositories/data_journalism/data_bit_1/va_pnnl_researched.xlsx'
wb = openpyxl.load_workbook(XLSX)
ws = wb.active

hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
V  = hdrs['verified']
SY = hdrs['source_year_found']
SF = hdrs['source_year_field']
NO = hdrs['verification_note']
NS = hdrs['new_source_url']

def write(row, verified, src_year, src_field, note, new_src=''):
    ws.cell(row, V).value  = verified
    ws.cell(row, SY).value = src_year
    ws.cell(row, SF).value = src_field
    ws.cell(row, NO).value = note
    if new_src:
        ws.cell(row, NS).value = new_src

# ── Meta Henrico Data Center (4 identical rows) ────────────────────────────
# Meta blog post confirmed: "Henrico County, We Are Online!" published Aug 5, 2020.
META_NOTE = (
    'Meta blog confirms Henrico Data Center went online August 5, 2020. '
    'Original cleanview.co source is JS-rendered (no parseable content).'
)
META_SRC = 'https://datacenters.atmeta.com/2020/08/henrico-county-we-are-online/'
for row in [2, 146, 207, 222]:
    write(row, 'Verified (new source)', '2020', 'Operational announcement',
          META_NOTE, META_SRC)

# ── Evocative IAD1 (row 16) ───────────────────────────────────────────────
# Original baxtel URL 404. New page at evocative-virginia-iad1 states:
# "constructed in 1982 and was fully renovated in 2008."
# Research year 2008, basis "construction completion" = renovation completion.
write(16,
      'Verified (new source)',
      '2008', 'Renovation year',
      'Building originally constructed 1982; fully renovated in 2008 per baxtel. '
      'Original source URL returned 404; new baxtel page found.',
      'https://baxtel.com/data-center/evocative-virginia-iad1')

# ── Flexential Richmond (row 29) ──────────────────────────────────────────
# Original URL 404. New baxtel page shows Year Built: 2011.
# Research says construction completion 2001 — mismatch.
write(29,
      'Mismatch',
      '2011', 'Year Built (new source)',
      'Original baxtel source URL returned 404. New baxtel page '
      '(flexential-richmond-1) shows Year Built: 2011. '
      'Research year 2001 (construction completion) does not match. '
      'Facility was formerly Peak 10; possible data discrepancy.',
      'https://baxtel.com/data-center/flexential-richmond-1')

# ── COPT Data Center rows (22, 210) ──────────────────────────────────────
# Source is a 2018 marketing PDF brochure (cdn.baxtel.com). Research year 2018
# (announcement year). datacenter.fyi shows COPT DC-6 actual operational date
# Aug 24, 2021 — a different metric from announcement year.
COPT_NOTE = (
    'Source is a 2018 marketing PDF brochure; year cannot be extracted from PDF. '
    'Research year 2018 is the announcement/marketing year. '
    'datacenter.fyi shows COPT DC-6 became operationally certified Aug 24, 2021 '
    '(different metric: permit operational date vs. project announcement).'
)
COPT_SRC = 'https://www.datacenter.fyi/public-record/copt-dc6-ec42737b'
write(22,  'Partial - see note', '2021', 'Actual Operational (datacenter.fyi)', COPT_NOTE, COPT_SRC)
write(210, 'Partial - see note', '2021', 'Actual Operational (datacenter.fyi)', COPT_NOTE, COPT_SRC)

# ── DBT Data Harrisonburg (row 164) ───────────────────────────────────────
# Original dbtdata.com URL 404. The Harrisonburg Cyber Integration Center:
# - DBT acquired it Oct 2013; sold to Harris 2010; repurchased 2013.
# - It was subsequently sold in Sept 2018 (divestiture).
# Research year 2018 (announcement year) likely refers to the 2018 sale.
# datacenter.fyi page for this facility has no operational date.
write(164,
      'Could not verify',
      '2018', 'Sale year (research basis)',
      'Original source (dbtdata.com) returned 404. '
      'DBT sold the Harrisonburg Cyber Integration Center in September 2018 '
      '(confirmed via datacenterknowledge.com). Research year 2018 (announcement year) '
      'likely refers to this divestiture, not an opening date. '
      'No opening/operational date found for this facility.',
      'https://www.datacenterknowledge.com/business/dbt-data-repurchases-virginia-site-from-harris')

# ── Amazon IAD128 (row 166) ───────────────────────────────────────────────
# Original source (IAD-124-127 page, Dec 2023) is for different facilities.
# Baxtel page for IAD-128 shows "Year Planned: 2026" — not 2024.
# Research says announcement year 2024; baxtel planned completion is 2026.
write(166,
      'Mismatch / Could not verify',
      '2026', 'Year Planned (baxtel)',
      'Original source covered IAD-124/125/126/127 (Dec 2023), not IAD-128. '
      'Baxtel page for IAD-128 shows "Year Planned: 2026" (under construction). '
      'Research year 2024 (announcement year) not confirmed; '
      'planned completion on source is 2026.',
      'https://baxtel.com/data-center/amazon-iad-128')

# ── Amazon IAD601 (row 208) ───────────────────────────────────────────────
# Original source (IAD-124-127) is wrong for this facility.
# No baxtel page found for IAD-601; IAD-600 exists but is different.
write(208,
      'Could not verify - source mismatch',
      None, None,
      'Original source covered IAD-124/125/126/127 (Dec 2023), not IAD-601. '
      'No dedicated baxtel or datacenter.fyi page found for IAD-601 specifically. '
      'Research year 2024 (announcement year) cannot be confirmed.',
      '')

# ── CoreSite Reston Campus VA2 (row 223) ──────────────────────────────────
# Original source: baxtel company page (not individual DC page).
# Baxtel individual page shows Year Built: 2014.
# Research says 2008, year opened — but 2008 was when CoreSite acquired the
# adjacent VA1 property (former AOL). VA2 itself was built in 2014.
write(223,
      'Mismatch',
      '2014', 'Year Built (new source)',
      'Source was a CoreSite company overview page, not the VA2 page. '
      'Baxtel VA2 page shows Year Built: 2014. '
      'Research year 2008 (year opened) likely refers to CoreSite acquiring '
      'the adjacent VA1 / former AOL campus in 2008, not VA2 itself.',
      'https://baxtel.com/data-center/coresite-va2')

# ── Microsoft planned DCs (rows 3, 306) ──────────────────────────────────
# Baxtel page says "Planned" with no Year Built/Planned field.
# False-positive 1982 came from nearby-facilities text — NOT the main facility.
MSFT_NOTE = (
    'Facility is planned/under construction. Baxtel page shows "Planned" status '
    'with no Year Built or Year Planned field. Research year 2027 '
    '(construction completion) is a projected date not explicitly confirmed on source. '
    'Any year extracted from this page reflects nearby facilities, not this one.'
)
write(3,   'Could not verify', None, None, MSFT_NOTE)
write(306, 'Could not verify', None, None, MSFT_NOTE)

# ── Google Bristow Campus (rows 253, 268, 310) ────────────────────────────
# Research year 2024 (announcement year). Baxtel shows Year Planned: 2026.
# March 2024: Google-linked company received approval for 181-acre Bristow campus.
# Source metric (Year Planned) differs from research metric (announcement year).
GOOGLE_NOTE = (
    'Research year 2024 is the announcement year (Google-linked company received '
    'approval for Bristow campus in March 2024 per press reports). '
    'Baxtel source shows "Year Planned: 2026" — planned completion year, '
    'a different metric. Announcement year 2024 is plausible but not directly '
    'stated on source page.'
)
for row in [253, 268, 310]:
    write(row, 'Partial - see note', '2026', 'Year Planned', GOOGLE_NOTE)

# ── Microsoft/future-facility false-positive fix ───────────────────────────
# Re-clear any leftover "1982" extractions from the first pass
for row in [3, 306]:
    ws.cell(row, SY).value = None
    ws.cell(row, SF).value = None

wb.save(XLSX)
print('Final update complete. Saved.')

# ── Summary ────────────────────────────────────────────────────────────────
counts = {}
for row in range(2, ws.max_row + 1):
    v = ws.cell(row, V).value or 'Unknown'
    counts[v] = counts.get(v, 0) + 1

print('\nVerification summary:')
for status, n in sorted(counts.items(), key=lambda x: -x[1]):
    print(f'  {n:3d}  {status}')
