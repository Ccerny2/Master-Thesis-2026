"""
Verifies research_year values in va_pnnl_researched.xlsx against their source URLs.
Adds columns: verified, source_year_found, source_year_field, verification_note, new_source_url
"""

import requests
from bs4 import BeautifulSoup
import re
import time
import openpyxl
from urllib.parse import urlparse

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/120.0.0.0 Safari/537.36'
    )
}
DELAY = 0.8  # seconds between requests to same domain
TIMEOUT = 20

# ── Extraction helpers ──────────────────────────────────────────────────────

def extract_baxtel(raw_html):
    """Return (year_str, field_label) from a baxtel page."""
    # 'Year Built:' appears in the main data-center card, before the nearby table
    m = re.search(r'Year Built:\s*(\d{4})', raw_html)
    if m:
        return m.group(1), 'Year Built'
    m = re.search(r'Year Opened:\s*(\d{4})', raw_html)
    if m:
        return m.group(1), 'Year Opened'
    return None, None


def extract_datacenter_fyi(raw_html):
    """Return (year_str, field_label) from a datacenter.fyi page."""
    # Operational Date label followed by a date value
    m = re.search(
        r'Operational Date</p>.*?(\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{1,2},?\s+(\d{4})|\b(\d{4})\b)',
        raw_html, re.DOTALL
    )
    if m:
        year = m.group(2) or m.group(3)
        return year, 'Operational Date'
    # Fallback: look for year in JSON-LD description
    m = re.search(r'"Operational Date[^"]*?(\d{4})', raw_html)
    if m:
        return m.group(1), 'Operational Date (JSON-LD)'
    return None, None


def extract_cleanview(raw_html, soup):
    """Return (year_str, field_label) from a cleanview.co page."""
    text = soup.get_text(' ', strip=True)
    m = re.search(r'(?:Year (?:Built|Opened)|Opened|Built)[:\s]+(\d{4})', text, re.IGNORECASE)
    if m:
        return m.group(1), m.group(0).split(':')[0].strip()
    m = re.search(r'(?:opened|launched|established)[^.]*?(\d{4})', text, re.IGNORECASE)
    if m:
        return m.group(1), 'text mention'
    return None, None


def extract_generic(raw_html, soup):
    text = soup.get_text(' ', strip=True)
    m = re.search(r'(?:Year (?:Built|Opened)|Opened|Built)[:\s]+(\d{4})', text, re.IGNORECASE)
    if m:
        return m.group(1), m.group(0).split(':')[0].strip()
    return None, None


def fetch_page(url, cache):
    if url in cache:
        return cache[url]
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        cache[url] = (r.status_code, r.text)
        return r.status_code, r.text
    except Exception as e:
        cache[url] = (0, str(e))
        return 0, str(e)


def extract_year(url, status, raw_html):
    if status == 0:
        return None, None, f'Request error: {raw_html[:120]}'
    if status != 200:
        return None, None, f'HTTP {status}'

    domain = urlparse(url).netloc
    soup = BeautifulSoup(raw_html, 'html.parser')

    if 'baxtel.com' in domain:
        year, field = extract_baxtel(raw_html)
    elif 'datacenter.fyi' in domain:
        year, field = extract_datacenter_fyi(raw_html)
    elif 'cleanview.co' in domain:
        year, field = extract_cleanview(raw_html, soup)
    else:
        year, field = extract_generic(raw_html, soup)

    if year:
        return year, field, None
    return None, None, 'No year field found on page'


# ── Main ────────────────────────────────────────────────────────────────────

XLSX = '/Users/elenamurray/Documents/Documents/Repositories/data_journalism/data_bit_1/va_pnnl_researched.xlsx'

wb = openpyxl.load_workbook(XLSX)
ws = wb.active

# Add or locate output columns (append to right)
existing_headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

def get_or_create_col(name):
    if name in existing_headers:
        return existing_headers[name]
    col = ws.max_column + 1
    ws.cell(1, col).value = name
    existing_headers[name] = col
    return col

col_verified     = get_or_create_col('verified')
col_src_year     = get_or_create_col('source_year_found')
col_src_field    = get_or_create_col('source_year_field')
col_note         = get_or_create_col('verification_note')
col_new_src      = get_or_create_col('new_source_url')

COL_RESEARCH_YEAR = 14  # research_year
COL_BASIS         = 15  # year_basis_used
COL_SOURCE_URL    = 16  # year_source_link

cache = {}
last_request_time = {}

total = ws.max_row - 1
print(f'Processing {total} rows...\n')

failed_rows = []  # rows where we couldn't verify — need follow-up search

for row in range(2, ws.max_row + 1):
    research_year = ws.cell(row, COL_RESEARCH_YEAR).value
    basis         = ws.cell(row, COL_BASIS).value
    url           = ws.cell(row, COL_SOURCE_URL).value
    operator      = ws.cell(row, 7).value
    name          = ws.cell(row, 9).value

    if not url:
        ws.cell(row, col_verified).value  = 'No source URL'
        ws.cell(row, col_note).value      = 'No source link in dataset'
        print(f'Row {row:3d}: no URL — skipping')
        continue

    # Rate-limit per domain
    domain = urlparse(url).netloc
    elapsed = time.time() - last_request_time.get(domain, 0)
    if elapsed < DELAY and url not in cache:
        time.sleep(DELAY - elapsed)

    status, raw = fetch_page(url, cache)
    if url not in cache or cache[url][0] != status:
        last_request_time[domain] = time.time()
    else:
        last_request_time[domain] = time.time()

    src_year, src_field, err = extract_year(url, status, raw)

    ws.cell(row, col_src_year).value  = src_year
    ws.cell(row, col_src_field).value = src_field

    if err:
        ws.cell(row, col_verified).value = 'Could not verify'
        ws.cell(row, col_note).value     = err
        failed_rows.append(row)
        print(f'Row {row:3d}: ERROR — {err[:60]}  | {name}')
    elif src_year == str(research_year):
        ws.cell(row, col_verified).value = 'Verified'
        ws.cell(row, col_note).value     = f'Source shows {src_field}: {src_year}'
        print(f'Row {row:3d}: ✓ {research_year}  | {name}')
    else:
        ws.cell(row, col_verified).value = 'Mismatch / Could not verify'
        ws.cell(row, col_note).value     = (
            f'Research year {research_year} ({basis}); '
            f'source shows {src_field}: {src_year}'
        ) if src_year else (
            f'Research year {research_year} ({basis}); no year found on source page'
        )
        failed_rows.append(row)
        print(f'Row {row:3d}: ✗ research={research_year} vs source={src_year}  | {name}')

    # Save progress every 50 rows
    if (row - 1) % 50 == 0:
        wb.save(XLSX)
        print(f'  [saved at row {row}]')

wb.save(XLSX)
print(f'\nDone. Saved. {len(failed_rows)} rows need follow-up: {failed_rows}')
